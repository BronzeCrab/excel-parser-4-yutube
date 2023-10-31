from datetime import date
from openpyxl import load_workbook

from models import Company, Foo, Bar, db


FILE_PATH = "companies_data.xlsx"


def create_db():
    try:
        # Connect to our database.
        db.connect()
        # Create the tables.
        db.create_tables([Company, Foo, Bar])
    except Exception as exp:
        print("ERROR: creating db, cause: {0}".format(exp))
        exit()


def load_data():
    try:
        wb = load_workbook(FILE_PATH)
    except Exception as exp:
        print("ERROR: opening file, cause: {0}".format(exp))
        exit()

    ws = wb.active

    cell_range = ws["A3":"F22"]

    for row in cell_range:
        for cell in row:
            if cell.column_letter == "A":
                day_num = cell.value
            elif cell.column_letter == "B":
                company, _ = Company.get_or_create(name=cell.value)
            elif cell.column_letter == "C":
                data1 = cell.value
            elif cell.column_letter == "D":
                Foo.create(
                    company_id=company.id,
                    data1=data1,
                    data2=cell.value,
                    date=date(1960, 1, day_num),
                )
            elif cell.column_letter == "E":
                data1 = cell.value
            elif cell.column_letter == "F":
                Bar.create(
                    company_id=company.id,
                    data1=data1,
                    data2=cell.value,
                    date=date(1960, 1, day_num),
                )


def main():
    create_db()
    load_data()


if __name__ == "__main__":
    main()
